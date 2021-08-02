from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

import workbook

wb = workbook.get_workbook()


def create_transaction_sheet():
    transaction = wb.create_sheet("Transactions")
    ColumnDimension(transaction, bestFit=True)
    transaction.append(
        (("NAV Date", "Transaction Type", "Units", "NAV", "Amount", "Total Units")))
    row = transaction.row_dimensions[transaction.max_row]
    row.font = Font(bold=True)


def update_transaction_sheet(summary_details, transactions):
    date_gross_list_of_tupple = []
    transaction = wb["Transactions"]
    transaction.append((summary_details[1],))
    for transaction_tupple in transactions:
        transaction.append((transaction_tupple))
        date_gross_list_of_tupple.append((transaction_tupple[0], transaction_tupple[4]))
    transaction.append(())
    transaction.append(
        ("Current NAV: ", summary_details[10], "Current Value: ", summary_details[6], "XIRR: ", summary_details[8]))
    transaction.cell(row=transaction.max_row, column=transaction.max_column).number_format = FORMAT_PERCENTAGE_00
    transaction.append(
        ("NAV Date: ", summary_details[9], "Cost Value: ", summary_details[2], "Redemptions : ", summary_details[3]))
    transaction.append(
        ("", "", "Switch In: ", summary_details[4], "Switch Out: ", summary_details[5]))
    transaction.append(())
    return date_gross_list_of_tupple


def finish_transaction_page():
    ws = wb["Transactions"]
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = FORMAT_NUMBER_00

    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[get_column_letter(i)].best_fit = False
        ws.column_dimensions[get_column_letter(i)].auto_size = True

        ColumnDimension(ws, index=letter, width=16)
