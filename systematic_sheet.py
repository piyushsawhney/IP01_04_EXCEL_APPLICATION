from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00, FORMAT_DATE_DDMMYY
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

import workbook

wb = workbook.get_workbook()


def create_systematic_sheet():
    systematic = wb.create_sheet("Systematic")
    ColumnDimension(systematic, bestFit=True)
    systematic.append(
        (("Type", "Status", "Folio No", "Scheme Name", "Amount", "Start Date", "End Date", "Cease Date",
          "Bank Name", "Bank Account No", "Target Scheme Name")))
    # row = systematic.row_dimensions[systematic.max_row]
    max_row = f"{systematic.max_row}:{systematic.max_row}"
    for cell in systematic[max_row]:
        cell.font = Font(bold=True)
    systematic.freeze_panes = "A2"


def update_systematic_sheet(systematic_details):
    systematic = wb["Systematic"]
    if systematic_details:
        for systematic_detail in systematic_details:
            systematic.append((systematic_detail))


def finish_systematic_page():
    ws = wb["Systematic"]
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = FORMAT_NUMBER_00
        ws.cell(row=row, column=6).number_format = FORMAT_DATE_DDMMYY
        ws.cell(row=row, column=7).number_format = FORMAT_DATE_DDMMYY
        ws.cell(row=row, column=8).number_format = FORMAT_DATE_DDMMYY

    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[get_column_letter(i)].best_fit = False
        ws.column_dimensions[get_column_letter(i)].auto_size = True

        ColumnDimension(ws, index=letter, width=16)
