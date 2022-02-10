import datetime

from openpyxl.drawing import image
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

import workbook

wb = workbook.get_workbook()


def create_summary_sheet(pan, investor_name=None):
    summary = wb.active
    summary.title = "Summary"
    ColumnDimension(summary, bestFit=True)
    img = image.Image('logo.png')
    img.height = 125
    img.width = 470
    img.anchor = 'C1'
    summary.add_image(img)
    summary.append(())
    summary.append(())
    summary.append(())
    summary.append(())
    summary.append(())
    summary.append(())
    summary.append(("Name", investor_name, "", "", "", "", "", str(datetime.date.today().strftime("%d-%B-%Y"))))
    summary.append(("PAN", pan))
    summary.append(())
    summary.append(
        ("Folio Number", "Scheme Name", "Cost Value", "Redemptions", "Switch In", "Switch Out", "Current Value",
         "Total Units", "Returns"))
    max_row = f"{summary.max_row}:{summary.max_row}"
    for cell in summary[max_row]:
        cell.font = Font(bold=True)
    summary.freeze_panes = "A11"


def update_summary_sheet(scheme_summary_details, file_name):
    file_name = f'{file_name.upper()}.xlsx'
    summary = wb["Summary"]
    summary.append((scheme_summary_details))
    summary.cell(row=summary.max_row,
                 column=2).value = f'=HYPERLINK("[{file_name}]Transactions!A{wb["Transactions"].max_row + 1}","{scheme_summary_details[1]}")'


def finish_summary_page(xirr):
    ws = wb["Summary"]
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    max_row = ws.max_row
    ws.cell(row=max_row + 2, column=3).value = f"=SUM(C5:C{max_row})"
    ws.cell(row=max_row + 2, column=4).value = f"=SUM(D5:D{max_row})"
    ws.cell(row=max_row + 2, column=5).value = f"=SUM(E5:E{max_row})"
    ws.cell(row=max_row + 2, column=6).value = f"=SUM(F5:F{max_row})"
    ws.cell(row=max_row + 2, column=7).value = f"=SUM(G5:G{max_row})"
    ws.cell(row=max_row + 2, column=9).value = xirr
    for row in range(5, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 3):
        ws.cell(row=row, column=6).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 3):
        ws.cell(row=row, column=7).number_format = FORMAT_NUMBER_00
    for row in range(5, ws.max_row + 3):
        ws.cell(row=row, column=9).number_format = FORMAT_PERCENTAGE_00

    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[get_column_letter(i)].best_fit = False
        ws.column_dimensions[get_column_letter(i)].auto_size = True

        ColumnDimension(ws, index=letter, width=16)
